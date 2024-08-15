import openpyxl
from openpyxl import Workbook
import os
import re
from PyPDF2 import PdfReader

class Bot:
    
    def action(self, execution=None):
        # List of PDF files to process
        pdf_files = [
            #r"C:\PDF\New Data\Structured_Doc_2.pdf", r"C:\PDF\New Data\Structured_Doc_3.pdf", 
            r"C:\PDF\New Data\Structured_Doc_4.pdf", r"C:\PDF\New Data\Structured_Doc_5.pdf", r"C:\PDF\New Data\Structured_Doc_6.pdf", r"C:\PDF\New Data\Structured_Doc_8.pdf", r"C:\PDF\New Data\Structured_Doc_9.pdf" # Replace with actual PDF paths
            # Add more PDF file paths here
        ]
        
        # Function to extract text from PDF
        def extract_text_from_pdf(pdf_path):
            with open(pdf_path, 'rb') as file:
                reader = PdfReader(file)
                text = ""
                for page_num in range(len(reader.pages)):
                    page = reader.pages[page_num]
                    text += page.extract_text()
                return text

        # Function to map data
        def map_invoice_data(text):
            invoice_data = {}

            # Adjusted regex patterns
            invoice_number_match = re.search(r'INV-\d+', text)
            order_number_match = re.search(r'Order  Number\s+\s+(\d+)', text)

            # Extracting dates
            invoice_date_match = re.search(r'Invoice\s+\s+Date\s+([A-Za-z]+) (\d{1,2}), (\d{4})', text)
            due_date_match = re.search(r'Due\s+Date\s+([A-Za-z]+) (\d{1,2}), (\d{4})', text)
            
            # Extracting 'From' and 'To' details
            from_details_match = re.search(r'From:\s+([\s\S]+?)\s+To:', text)
            to_details_match = re.search(r'To:\s+([\s\S]+?)\s+(?:Hrs/Qty|$)', text)
            
            # Extracting email addresses
            from_email_match = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', from_details_match.group(1)) if from_details_match else None
            to_email_match = re.search(r'([a-zA-Z0-9._%+-]+@\s+[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', to_details_match.group(1)) if to_details_match else None
            # Remove emails from 'From' and 'To' addresses
            from_address = re.sub(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', '', from_details_match.group(1)).strip() if from_details_match else 'N/A'
            to_address = re.sub(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', '', to_details_match.group(1)).strip() if to_details_match else 'N/A'

            # Extracting service details
            service_details_pattern = re.compile(r'(\d+\.\d{2})\s+([\w\s]+?)\s+(\d+\.\d{2})\s+(\d+\.\d{2}%)\s+\$(\d+\.\d{2})')
            service_details_matches = service_details_pattern.findall(text)

            service_entries = [{'Quantity/Hours': match[0], 'Description': match[1].strip(),
                                'Unit Price': match[2], 'Tax Rate': match[3], 'Amount': match[4]}
                               for match in service_details_matches]

            # If only one service entry, ensure it is placed in Service 1
            if len(service_entries) == 1:
                service_entries += [{'Quantity/Hours': 'N/A', 'Description': 'N/A',
                                     'Unit Price': 'N/A', 'Tax Rate': 'N/A', 'Amount': 'N/A'}] * 2
            elif len(service_entries) == 2:
                service_entries.append({'Quantity/Hours': 'N/A', 'Description': 'N/A',
                                        'Unit Price': 'N/A', 'Tax Rate': 'N/A', 'Amount': 'N/A'})

            # Extracting financial details
            sub_total_match = re.search(r'Sub Total\s*\$([\d,]+\.\d+)', text)
            tax_match = re.search(r'Tax\s*\$([\d,]+\.\d+)', text)
            total_due_match = re.search(r'Total\s+Due\s+\$([\d,]+\.\d+)', text)

            # Storing extracted data
            invoice_data['Invoice Number'] = invoice_number_match.group(0) if invoice_number_match else 'N/A'
            invoice_data['Order Number'] = order_number_match.group(1) if order_number_match else 'N/A'
            invoice_data['Invoice Date'] = f"{invoice_date_match.group(1)} {invoice_date_match.group(2)}, {invoice_date_match.group(3)}" if invoice_date_match else 'N/A'
            invoice_data['Due Date'] = f"{due_date_match.group(1)} {due_date_match.group(2)}, {due_date_match.group(3)}" if due_date_match else 'N/A'
            invoice_data['From'] = from_address
            invoice_data['From Email'] = from_email_match.group(0) if from_email_match else 'N/A'
            invoice_data['To'] = to_address
            invoice_data['To Email'] = to_email_match.group(0) if to_email_match else 'N/A'
            invoice_data['Services'] = service_entries
            invoice_data['Sub Total'] = sub_total_match.group(1) if sub_total_match else 'N/A'
            invoice_data['Tax'] = tax_match.group(1) if tax_match else 'N/A'
            invoice_data['Total Due'] = total_due_match.group(1) if total_due_match else 'N/A'
            
            return invoice_data

        # Function to save mapped data to an Excel file
        def save_mapped_data_to_excel(data, file_path):
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
            else:
                workbook = Workbook()
            
            sheet = workbook.active
            sheet.title = "Invoice Data"
            
            # If the sheet is empty, write headers
            if sheet.max_row == 1 and sheet.max_column == 1:
                headers = ['Invoice Number', 'Order Number', 'Invoice Date', 'Due Date', 
                           'From', 'From Email', 'To', 'To Email', 
                           'Service 1 - Quantity/Hours', 'Service 1 - Description', 'Service 1 - Unit Price', 'Service 1 - Tax Rate', 'Service 1 - Amount',
                           'Service 2 - Quantity/Hours', 'Service 2 - Description', 'Service 2 - Unit Price', 'Service 2 - Tax Rate', 'Service 2 - Amount',
                           'Service 3 - Quantity/Hours', 'Service 3 - Description', 'Service 3 - Unit Price', 'Service 3 - Tax Rate', 'Service 3 - Amount',
                           'Sub Total', 'Tax', 'Total Due']
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)
            
            # Append data to the next available row
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=data['Invoice Number'])
            sheet.cell(row=next_row, column=2, value=data['Order Number'])
            sheet.cell(row=next_row, column=3, value=data['Invoice Date'])
            sheet.cell(row=next_row, column=4, value=data['Due Date'])
            sheet.cell(row=next_row, column=5, value=data['From'])
            sheet.cell(row=next_row, column=6, value=data['From Email'])
            sheet.cell(row=next_row, column=7, value=data['To'])
            sheet.cell(row=next_row, column=8, value=data['To Email'])
            
            column_start = 9
            for service_entry in data['Services']:
                sheet.cell(row=next_row, column=column_start, value=service_entry['Quantity/Hours'])
                sheet.cell(row=next_row, column=column_start + 1, value=service_entry['Description'])
                sheet.cell(row=next_row, column=column_start + 2, value=service_entry['Unit Price'])
                sheet.cell(row=next_row, column=column_start + 3, value=service_entry['Tax Rate'])
                sheet.cell(row=next_row, column=column_start + 4, value=service_entry['Amount'])
                column_start += 5
            
            sheet.cell(row=next_row, column=column_start, value=data['Sub Total'])
            sheet.cell(row=next_row, column=column_start + 1, value=data['Tax'])
            sheet.cell(row=next_row, column=column_start + 2, value=data['Total Due'])
            
            # Save the workbook
            workbook.save(file_path)

        # Define the output Excel file path
        output_file_path = 'C:\PDF\ModifiedExcel.xlsx'  # Use the path of the uploaded Excel file

        # Process each PDF file
        for pdf_path in pdf_files:
            # Extract text from the PDF
            extracted_text = extract_text_from_pdf(pdf_path)

            # Map the data
            mapped_data = map_invoice_data(extracted_text)

            # Save the mapped data to an Excel file
            save_mapped_data_to_excel(mapped_data, output_file_path)

        print(f"Extracted and mapped invoice data has been saved to {output_file_path}")

    def not_found(self, label):
        print(f"Element not found: {label}")

if __name__ == "__main__":
    Bot().action()